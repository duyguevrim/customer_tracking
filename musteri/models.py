from company import db
from openpyxl import load_workbook
from company.models import User
from datetime import datetime


class Musteri(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    isim = db.Column(db.String, nullable=False)
    telefon = db.Column(db.String, nullable=False)
    bakiye = db.Column(db.Float, nullable=False)
    guncel_bakiye = db.Column(db.Float, default=0.0)
    tahsilat = db.Column(db.Float, default=0.0)
    mail = db.Column(db.String)
    parasut_no = db.Column(db.String, nullable=False)
    parasut_odeme_linki = db.Column(db.String)
    tahsilatlar = db.relationship('Tahsilat', backref='musteri', lazy=True)
    odemeler = db.relationship('Odeme', backref='musteri', lazy=True)
    indirimler = db.relationship('Indirim', backref='musteri', lazy='dynamic')
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)


class Tahsilat(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tutar = db.Column(db.Float, nullable=False)
    durum = db.Column(db.String, nullable=False, default="Onay Bekliyor")
    dekont = db.Column(db.String, nullable=True)
    aciklama = db.Column(db.Text, nullable=True)
    shortcode = db.Column(db.String, nullable=False)
    create_date = db.Column(db.Date, nullable=False, default=datetime.now())
    parasut_id = db.Column(db.String)
    musteri_id = db.Column(db.Integer, db.ForeignKey('musteri.id'), nullable=False)


class Odeme(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tutar = db.Column(db.Float, nullable=False)
    create_date = db.Column(db.Date, nullable=False)
    musteri_id = db.Column(db.Integer, db.ForeignKey('musteri.id'), nullable=False)


class Indirim(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    indirim_tutari = db.Column(db.Float, nullable=False)
    musteri_id = db.Column(db.Integer, db.ForeignKey('musteri.id'))
    create_date = db.Column(db.Date)








#
# ayge_excel = load_workbook("ids/zeynep-sari.xlsx")
# ayge_sheet = ayge_excel.active
#
# index = 2
#
# user1 = User.query.get(19)
#
# while index <= ayge_sheet.max_row:
#     musteri_adi = ayge_sheet["A" + str(index)].value
#     mail = ayge_sheet["B" + str(index)].value
#     telefon = ayge_sheet["C" + str(index)].value
#     bakiye = ayge_sheet["D" + str(index)].value
#     parasut_no = ayge_sheet["E" + str(index)].value
#     parasut_link = ayge_sheet["F" + str(index)].value
#     musteri = Musteri(isim=musteri_adi, telefon=telefon, bakiye=bakiye, mail=mail, parasut_no=parasut_no, parasut_odeme_linki=parasut_link, user_id=user1.id)
#     db.session.add(musteri)
#     db.session.commit()
#     index += 1
